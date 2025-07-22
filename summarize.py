#!/usr/bin/env python3
# encoding: utf-8

r"""
# summarize: Convert Screen Shots into a Readable Text and create a summary; also works with PDFs.

This script provides functionality to convert screenshots and PDFs into a readable text format,
generate summaries, and create styled PDF outputs. It can operate in two modes:

1. **Image Mode**: Process PNG/JPG/JPEG screenshots in the current directory
2. **PDF Mode**: Recursively analyze PDFs in subdirectories with financial summary

# Installation

To install the required libraries, run the following command:

```bash
pip install -r requirements.txt
```

You may also need to

```bash
python -m spacy download en_core_web_sm
```


# Configuration

Before using the script, you need to configure your AI settings. You can do
this by running the following command:

```bash
./summarize.py config
```

Then you should define your default AI provider

```bash
./summarize.py config -d
```

# Usage

To get help about the script, call it with the `--help` option:

```bash
./summarize.py --help
```

## Automatic Mode Detection

The script automatically detects which mode to use:

```bash
./summarize.py
```

- If PNG/JPG/JPEG files are found in the current directory → Image mode
- If PDFs are found in subdirectories → PDF recursive analysis mode

## Image Mode: Summarize screenshots

```bash
./summarize.py
```

This will process PNG, JPG, and JPEG files in the current directory and create a styled PDF summary.

## PDF Mode: Recursive financial analysis

```bash
./summarize.py
```

When PDFs are detected in subdirectories, the script will:
- Recursively scan all subdirectories for PDF files
- Extract text and amounts from each PDF
- Generate AI summaries for each directory
- Create aggregate summaries for parent directories
- Output three files at the root level:
  - `summary_TIMESTAMP.json` - Complete analysis data
  - `summary_TIMESTAMP.xlsx` - Excel workbook with financial summary
  - `summary_TIMESTAMP.pdf` - Formatted PDF report

### Features:
- **Multi-currency support**: Automatically detects and tracks any valid ISO currency
- **Smart amount extraction**: Finds total amounts in receipts across multiple languages
- **Parallel processing**: Uses multiple threads for faster analysis
- **Rate limit handling**: Gracefully handles API rate limiting
- **Non-receipt PDFs**: Processes all PDFs, even non-financial documents

### Output Format:
The Excel summary includes:
- Directory path
- Document type (PDFs or Aggregate)
- Number of documents
- Processing date
- Separate columns for each currency found
- Total row with sums (using SUBTOTAL for filtering support)

## Summarize specific files

```bash
./summarize.py file1.png file2.pdf
```

# Documentation

To generate the documentation for the script, run the following command:

```bash
./summarize.py doc
```

# License

This script is released under the [WTFPL License](https://en.wikipedia.org/wiki/WTFPL).
"""

# Standard library imports
import contextlib
import glob
import io
import json
import logging
import multiprocessing
import os
import re
import shutil
import sys
import tempfile
import warnings
from collections import defaultdict
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict, Tuple

# Third-party imports
import cssutils
import pymupdf as fitz
import img2pdf
import markdown
import nltk
import numpy as np
import pandas as pd
import pytesseract
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table as ExcelTable, TableStyleInfo
import spacy
import traceback
import typer
from bs4 import BeautifulSoup
from InquirerPy import inquirer
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize, sent_tokenize
from pdf2image import convert_from_path
from PIL import Image, ImageEnhance, ImageChops, UnidentifiedImageError
from PyPDF2 import PdfMerger, PdfReader, PdfWriter
from reportlab.lib import colors
from reportlab.lib.colors import HexColor, Color, black, white
from reportlab.lib.enums import TA_JUSTIFY, TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, ListFlowable, ListItem, PageBreak, Table, TableStyle
from rich import print
from rich.console import Console
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn, TimeElapsedColumn
from skimage.metrics import structural_similarity as ssim
from transformers import pipeline

# Local imports
from client import AIClient, GitConfig, DocGenerator

# Suppress warnings
warnings.filterwarnings("ignore")

# Initialize console
console = Console()

# Initialize Typer app
app = typer.Typer(
    add_completion=False,
    rich_markup_mode="rich",
    no_args_is_help=True,
    help="summarize: Convert Screen Shots into a Readable Text and create a summary; also works with PDFs.",
    epilog="To get help about the script, call it with the --help option."
)



class SuppressOutput:
    """
    A context manager for suppressing stdout and stderr output.
    """
    def __enter__(self):
        self.stdout = os.dup(1)
        self.stderr = os.dup(2)
        devnull = os.open(os.devnull, os.O_WRONLY)
        os.dup2(devnull, 1)
        os.dup2(devnull, 2)
        os.close(devnull)
        logging.disable(logging.CRITICAL)
        return self

    def __exit__(self, *args):
        os.dup2(self.stdout, 1)
        os.dup2(self.stderr, 2)
        os.close(self.stdout)
        os.close(self.stderr)
        logging.disable(logging.NOTSET)

# Suppress specific library logging
transformers_logger = logging.getLogger('transformers')
transformers_logger.setLevel(logging.CRITICAL)

# Download necessary NLTK data
with SuppressOutput():
    nltk.download('punkt', quiet=True)
    nltk.download('stopwords', quiet=True)


def find_pdfs_recursively(directory: Path) -> Dict[Path, List[Path]]:
    """
    Find all PDF files recursively in a directory and group them by parent directory.
    Excludes generated summary PDFs to avoid double processing.
    
    Args:
        directory: Root directory to search
        
    Returns:
        Dictionary mapping directory paths to lists of PDF files in that directory
    """
    pdf_map = {}
    
    # Pattern to match our generated summary PDFs
    summary_pattern = re.compile(r'^summary_\d{8}_\d{6}\.pdf$')
    
    for root, dirs, files in os.walk(directory):
        root_path = Path(root)
        pdf_files = []
        
        for f in files:
            if f.lower().endswith('.pdf'):
                # Exclude our generated summary PDFs
                if not summary_pattern.match(f):
                    pdf_files.append(root_path / f)
        
        if pdf_files:
            pdf_map[root_path] = pdf_files
            
    return pdf_map


def extract_text_from_pdf_simple(pdf_path: Path) -> Dict[str, any]:
    """
    Extract all text from a PDF file (without OCR - assumes text is already embedded).
    
    Args:
        pdf_path: Path to the PDF file
        
    Returns:
        Dictionary containing extracted text and metadata
    """
    try:
        doc = fitz.open(str(pdf_path))
        
        full_text = []
        page_texts = []
        
        for page_num, page in enumerate(doc, 1):
            # Extract text from the page
            text = page.get_text()
            
            # Clean up the text
            text = text.strip()
            
            # Also check for text in annotations
            annot_texts = []
            for annot in page.annots():
                if annot:
                    content = annot.info.get('content', '')
                    if content and content.strip():
                        annot_texts.append(content.strip())
            
            # Combine regular text and annotation text
            if annot_texts and not text:
                text = "Annotations: " + "; ".join(annot_texts)
            elif annot_texts and text:
                text = text + "\nAnnotations: " + "; ".join(annot_texts)
            
            # Check if page has images but no text (likely needs OCR)
            if not text and page.get_images():
                text = "[Page contains images but no extractable text - OCR may be needed]"
            
            if text:
                page_texts.append({
                    'page': page_num,
                    'text': text
                })
                full_text.append(f"--- Page {page_num} ---\n{text}")
        
        doc.close()
        
        # Get file metadata
        stat = pdf_path.stat()
        
        return {
            'filename': pdf_path.name,
            'path': str(pdf_path),
            'pages': len(page_texts),
            'modified': datetime.fromtimestamp(stat.st_mtime).isoformat(),
            'size_bytes': stat.st_size,
            'full_text': '\n\n'.join(full_text),
            'page_texts': page_texts
        }
        
    except Exception as e:
        console.print(f"[red]Error extracting text from {pdf_path}: {e}[/red]")
        return {
            'filename': pdf_path.name,
            'path': str(pdf_path),
            'error': str(e)
        }


def extract_amounts_from_text(text: str) -> List[Tuple[float, str]]:
    """
    Extract monetary amounts and their currencies from text.
    Focus on total amounts in summaries.
    
    Returns:
        List of (amount, currency) tuples
    """
    # Common valid currency codes (ISO 4217)
    VALID_CURRENCIES = {
        'USD', 'EUR', 'GBP', 'CHF', 'JPY', 'AUD', 'CAD', 'CNY', 'SEK', 'NOK',
        'DKK', 'PLN', 'CZK', 'HUF', 'RON', 'BGN', 'HRK', 'RUB', 'TRY', 'BRL',
        'MXN', 'ARS', 'CLP', 'COP', 'PEN', 'UYU', 'INR', 'IDR', 'KRW', 'MYR',
        'PHP', 'SGD', 'THB', 'VND', 'ZAR', 'AED', 'SAR', 'ILS', 'NZD', 'TWD',
        'HKD', 'ISK', 'RSD', 'UAH', 'KZT', 'GEL', 'AMD', 'AZN', 'BYN', 'MDL',
        'MKD', 'ALL', 'BAM', 'EGP', 'MAD', 'TND', 'JOD', 'KWD', 'LBP', 'DZD'
    }
    
    amounts = []
    found_totals = set()  # Track found amounts to avoid duplicates
    
    # If the text indicates no extractable content, return empty
    if '[Page contains images but no extractable text' in text or 'OCR may be needed' in text:
        return []
    
    # Currency pattern - matches any 3-letter uppercase currency code
    currency_pattern = r'[A-Z]{3}'
    
    # Priority patterns - look for total amounts first
    # Updated patterns to handle thousands separators (1,234.56 or 1'234.56 or 1 234,56)
    priority_patterns = [
        # Total patterns with various formats - now handles thousands separators
        rf'Total amount:\s*({currency_pattern})\s*[:.]?\s*([\d,\'\s]+(?:\.\d+)?)',
        rf'Total amount:\s*([\d,\'\s]+(?:\.\d+)?)\s*({currency_pattern})',
        rf'Total:\s*({currency_pattern})\s*[:.]?\s*([\d,\'\s]+(?:\.\d+)?)',
        rf'Total:\s*([\d,\'\s]+(?:\.\d+)?)\s*({currency_pattern})',
        rf'Total spending[^:]*:\s*({currency_pattern})\s*[:.]?\s*([\d,\'\s]+(?:\.\d+)?)',
        rf'Total spending[^:]*:\s*([\d,\'\s]+(?:\.\d+)?)\s*({currency_pattern})',
        # French/German variants
        rf'Montant total:\s*({currency_pattern})\s*[:.]?\s*([\d,\'\s]+(?:\.\d+)?)',
        rf'Montant total:\s*([\d,\'\s]+(?:\.\d+)?)\s*({currency_pattern})',
        rf'Montant:\s*({currency_pattern})\s*[:.]?\s*([\d,\'\s]+(?:\.\d+)?)',
        rf'Montant:\s*([\d,\'\s]+(?:\.\d+)?)\s*({currency_pattern})',
        rf'Betrag:\s*({currency_pattern})\s*[:.]?\s*([\d,\'\s]+(?:\.\d+)?)',
        rf'Betrag:\s*([\d,\'\s]+(?:\.\d+)?)\s*({currency_pattern})',
    ]
    
    # Look for priority patterns first
    for pattern in priority_patterns:
        matches = re.finditer(pattern, text, re.IGNORECASE | re.MULTILINE)
        for match in matches:
            # Determine which group contains the currency and which contains the amount
            groups = match.groups()
            currency = None
            amount_str = None
            
            for group in groups:
                if group and re.match(r'^[A-Z]{3}$', group.upper()):
                    currency = group.upper()
                elif group:
                    amount_str = group
            
            if currency and amount_str:
                # Validate currency code
                if currency not in VALID_CURRENCIES:
                    continue
                    
                # Remove thousands separators (comma, apostrophe, space)
                amount_str = amount_str.replace(',', '').replace("'", '').replace(' ', '')
                try:
                    amount = float(amount_str)
                    
                    # Validate amount is reasonable
                    if amount < 0.01 or amount > 999999999:  # Less than 1 cent or more than 999 million
                        continue
                        
                    key = (amount, currency)
                    if key not in found_totals:
                        found_totals.add(key)
                        amounts.append(key)
                except ValueError:
                    continue
    
    # If we found totals, return them all (let the caller decide which to use)
    if amounts:
        return amounts
    
    # If no totals found, look for any amount mentions (less reliable)
    fallback_patterns = [
        rf'({currency_pattern})\s*[:.]?\s*(\d+(?:[.,]\d+)?)',
        rf'(\d+(?:[.,]\d+)?)\s*({currency_pattern})',
    ]
    
    for pattern in fallback_patterns:
            matches = re.finditer(pattern, text, re.IGNORECASE)
            for match in matches:
                # Determine which group contains the currency and which contains the amount
                groups = match.groups()
                currency = None
                amount_str = None
                
                for group in groups:
                    if group and re.match(r'^[A-Z]{3}$', group.upper()):
                        currency = group.upper()
                    elif group:
                        amount_str = group
                
                if currency and amount_str:
                    # Validate currency code
                    if currency not in VALID_CURRENCIES:
                        continue
                        
                    amount_str = amount_str.replace(',', '.')
                    try:
                        amount = float(amount_str)
                        
                        # Validate amount is reasonable
                        if amount < 0.01 or amount > 999999999:
                            continue
                        
                        key = (amount, currency)
                        if key not in found_totals:
                            found_totals.add(key)
                            amounts.append(key)
                    except ValueError:
                        continue
    
    return amounts


def create_pdf_summary(texts: List[Dict], directory: Path) -> str:
    """
    Create a summary of extracted PDF texts using AI.
    
    Args:
        texts: List of extracted text dictionaries
        directory: Directory being processed
        
    Returns:
        Summary text
    """
    # Prepare content for summarization
    content_parts = []
    
    for doc in texts:
        if 'error' not in doc:
            content_parts.append(f"File: {doc['filename']}")
            content_parts.append(f"Modified: {doc['modified']}")
            content_parts.append(f"Full text:\n{doc['full_text']}")
            content_parts.append("\n---\n")
    
    combined_text = '\n'.join(content_parts)
    
    # Use AI to generate summary
    ai_client = AIClient(config_provider=GitConfig())
    
    prompt = f"""
    Analyze the following receipt/document texts extracted from PDF files in {directory.name}.
    Create a structured summary that includes:
    
    1. Overview of all documents found
    2. For receipts: extract key information like:
       - Date (look for date formats, dates, datum, data)
       - Vendor/Company name
       - Total amount (look for: total, montant, betrag, importo, amount, CHF, EUR, USD, summe, somme)
       - Items purchased (if visible)
       - Payment method (credit card, debit, cash, bar, carte, karte)
    3. Total spending across all receipts (if applicable)
    4. Any patterns or notable observations
    
    IMPORTANT: 
    - The receipts may be in multiple languages (English, French, German, Italian).
    - Look for amount indicators like: Total, Montant, Betrag, Importo, Summe, Somme, CHF, EUR, USD
    - ALWAYS include a clear total line at the end in this format: "Total: CHF 123.45" (or EUR, USD, etc.)
    - If there are multiple currencies, list each total separately like:
      Total: CHF 100.00
      Total: EUR 50.00
    
    Format the output as a clear, structured text that can be used for later analysis.
    
    Documents to analyze:
    
    {combined_text}
    """
    
    try:
        response = ai_client.prompt(prompt, tokens=1500)
        return response
    except Exception as e:
        error_msg = str(e)
        
        # Check for specific error codes
        if "529" in error_msg:
            console.print(f"[red]⚠️  Rate limit exceeded - API is temporarily unavailable[/red]")
            console.print(f"[yellow]   The AI service is currently overloaded. Please try again in a few minutes.[/yellow]")
            return "Summary unavailable due to rate limiting. The API service is temporarily overloaded."
        elif "401" in error_msg:
            console.print(f"[red]Authentication error - please check your API key[/red]")
            return "Summary unavailable due to authentication error."
        elif "500" in error_msg or "502" in error_msg or "503" in error_msg:
            console.print(f"[red]Server error - the AI service is experiencing issues[/red]")
            return "Summary unavailable due to server error."
        else:
            console.print(f"[red]Error generating AI summary: {e}[/red]")
            return f"Error generating summary: {e}"


def process_pdf_directory(directory: Path, show_progress: bool = True) -> Dict:
    """
    Process all PDFs in a directory and create a summary file.
    
    Args:
        directory: Directory to process
        output_format: Output format ('txt' or 'json')
        show_progress: Whether to show progress messages
        
    Returns:
        Processing results
    """
    pdf_files = [f for f in directory.iterdir() if f.suffix.lower() == '.pdf']
    
    if not pdf_files:
        return {'directory': str(directory), 'pdfs': 0, 'status': 'no_pdfs'}
    
    if show_progress:
        console.print(f"[blue]Processing {len(pdf_files)} PDFs in {directory}[/blue]")
    
    # Extract text from all PDFs
    extracted_texts = []
    for pdf_file in pdf_files:
        if show_progress:
            console.print(f"  Extracting: {pdf_file.name}")
        extracted = extract_text_from_pdf_simple(pdf_file)
        extracted_texts.append(extracted)
    
    # Create summary
    summary = create_pdf_summary(extracted_texts, directory)
    
    # Extract amounts from the summary for structured data
    summary_amounts = extract_amounts_from_text(summary)
    
    # If no amounts found in summary, try to extract from individual documents
    if not summary_amounts:
        all_doc_amounts = []
        for doc in extracted_texts:
            if 'error' not in doc and 'full_text' in doc:
                doc_amounts = extract_amounts_from_text(doc['full_text'])
                all_doc_amounts.extend(doc_amounts)
        
        # Aggregate by currency
        currency_totals = defaultdict(float)
        for amount, currency in all_doc_amounts:
            currency_totals[currency] += amount
        
        summary_amounts = [(total, currency) for currency, total in currency_totals.items()]
    else:
        # If we found multiple amounts for the same currency, use the largest (likely the total)
        currency_amounts = defaultdict(list)
        for amount, currency in summary_amounts:
            currency_amounts[currency].append(amount)
        
        # For each currency, take the maximum amount (which should be the total)
        summary_amounts = [(max(amounts), currency) for currency, amounts in currency_amounts.items()]
    
    # Return all data in memory instead of writing files
    return {
        'directory': str(directory),
        'processed_date': datetime.now().isoformat(),
        'documents': extracted_texts,
        'summary': summary,
        'extracted_amounts': [{'amount': amt, 'currency': curr} for amt, curr in summary_amounts],
        'total_files': len(pdf_files),
        'status': 'success'
    }


def create_aggregate_summary(directory: Path, subdirectory_results: List[Dict]) -> Dict:
    """
    Create an aggregate summary for a parent directory based on subdirectory results.
    Returns both the summary text and structured data.
    """
    # Process all subdirectory summaries from in-memory data
    all_amounts = defaultdict(list)  # currency -> list of amounts
    total_receipts = 0
    subdirs_with_pdfs = []
    all_documents = []
    
    for result in subdirectory_results:
        if result.get('status') == 'success' and result.get('total_files', 0) > 0:
            dir_path = Path(result['directory'])
            subdirs_with_pdfs.append(dir_path.name)
            
            # Extract amounts from the result data
            if 'extracted_amounts' in result:
                for amt_data in result['extracted_amounts']:
                    all_amounts[amt_data['currency']].append(amt_data['amount'])
            
            # Count receipts/documents
            total_receipts += result.get('total_files', 0)
            
            # Collect all documents for reference
            all_documents.extend(result.get('documents', []))
    
    # Calculate totals by currency
    currency_totals = {}
    
    for currency, amounts in all_amounts.items():
        currency_total = sum(amounts)
        currency_totals[currency] = currency_total
    
    # Generate AI summary
    ai_client = AIClient(config_provider=GitConfig())
    
    prompt = f"""
    Create an aggregate summary for the directory '{directory.name}' which contains the following subdirectories with receipts/documents:
    
    Subdirectories analyzed: {', '.join(subdirs_with_pdfs)}
    Total documents found: {total_receipts}
    
    Currency totals found:
    {chr(10).join(f'- {currency}: {total:.2f}' for currency, total in currency_totals.items())}
    
    Create a concise summary that:
    1. Summarizes what types of expenses/documents were found
    2. Lists the total spending by currency
    3. Notes any patterns or interesting observations across the subdirectories
    
    Keep it brief and focused on the financial overview.
    """
    
    try:
        response = ai_client.prompt(prompt, tokens=800)
        return {
            'summary_text': response,
            'extracted_amounts': [{'amount': total, 'currency': curr} for curr, total in currency_totals.items()],
            'subdirectories': subdirs_with_pdfs,
            'total_documents': total_receipts
        }
    except Exception as e:
        # Fallback summary if AI fails
        summary = f"Aggregate Summary for {directory.name}\n\n"
        summary += f"Subdirectories analyzed: {len(subdirs_with_pdfs)}\n"
        summary += f"Total documents: {total_receipts}\n\n"
        summary += "Currency totals:\n"
        for currency, total in currency_totals.items():
            summary += f"- {currency}: {total:.2f}\n"
        return {
            'summary_text': summary,
            'extracted_amounts': [{'amount': total, 'currency': curr} for curr, total in currency_totals.items()],
            'subdirectories': subdirs_with_pdfs,
            'total_documents': total_receipts
        }


def cleanup_intermediate_files(directory: Path, timestamp: str, keep_root: bool = True):
    """
    Remove all intermediate summary files recursively, optionally keeping root files.
    
    Args:
        directory: Root directory
        timestamp: Run timestamp to identify files from this run
        keep_root: If True, keep files in the root directory
    """
    # Patterns for files to clean up
    patterns = [
        f'summary_{timestamp}.json',
        f'summary_{timestamp}.txt',
        f'summary_{timestamp}.xlsx'
    ]
    
    for pattern in patterns:
        for file in directory.rglob(pattern):
            # Skip root directory files if keep_root is True
            if keep_root and file.parent == directory:
                continue
                
            try:
                file.unlink()
                console.print(f"[yellow]Cleaned up {file.relative_to(directory)}[/yellow]")
            except Exception as e:
                console.print(f"[red]Error removing {file}: {e}[/red]")


def create_master_json(directory: Path, all_data: Dict, timestamp: str) -> Path:
    """
    Create a master JSON file with all data from the analysis.
    
    Args:
        directory: Root directory for the analysis
        all_data: Complete in-memory data structure
        timestamp: Run timestamp
        
    Returns:
        Path to the created master JSON file
    """
    master_data = {
        'analysis_date': datetime.now().isoformat(),
        'root_directory': str(directory),
        'directories': all_data
    }
    
    # Save master JSON
    master_json = directory / f"summary_{timestamp}.json"
    with open(master_json, 'w', encoding='utf-8') as f:
        json.dump(master_data, f, indent=2, ensure_ascii=False)
    
    return master_json


def create_summary_pdf_from_data(all_data: Dict, output_path: Path, css_file: str = None) -> None:
    """
    Create a styled PDF summary from the analysis data.
    
    Args:
        all_data: Complete in-memory data structure
        output_path: Path for the output PDF
        css_file: Optional CSS file for styling
    """
    # Create markdown content with meaningful summary text
    md_parts = [
        "# PDF Analysis Summary",
        f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        ""
    ]
    
    # Track totals and details for summary
    total_docs = 0
    currency_totals = defaultdict(float)
    directories_processed = []
    aggregate_dirs = []
    
    # Process data to gather information
    for rel_path, data in sorted(all_data.items()):
        if data.get('status') == 'success':
            if data.get('aggregate', False):
                aggregate_dirs.append(rel_path)
                doc_count = data.get('total_documents', 0)
            else:
                directories_processed.append(rel_path)
                doc_count = data.get('total_files', 0)
            
            total_docs += doc_count
            
            # Collect currency totals
            if data.get('extracted_amounts'):
                for amt in data['extracted_amounts']:
                    currency_totals[amt['currency']] += amt['amount']
    
    # Create summary text
    md_parts.append("## Analysis Overview")
    md_parts.append("")
    md_parts.append(f"This analysis was completed on **{datetime.now().strftime('%B %d, %Y at %H:%M')}**.")
    md_parts.append("")
    
    if directories_processed:
        md_parts.append("### Directories Analyzed")
        md_parts.append("")
        md_parts.append(f"Processed **{total_docs} PDF documents** across **{len(directories_processed)} directories**:")
        md_parts.append("")
        for dir_path in directories_processed:
            md_parts.append(f"- {dir_path}")
        md_parts.append("")
    
    if aggregate_dirs:
        md_parts.append("### Aggregate Summaries")
        md_parts.append("")
        md_parts.append("The following aggregate summaries were created:")
        md_parts.append("")
        for agg_dir in aggregate_dirs:
            md_parts.append(f"- {agg_dir}")
        md_parts.append("")
    
    md_parts.append("### Financial Summary")
    md_parts.append("")
    
    if currency_totals:
        md_parts.append("**Amounts by Currency:**")
        md_parts.append("")
        for currency, amount in sorted(currency_totals.items()):
            if amount > 0:  # Only show currencies with amounts
                md_parts.append(f"- **{currency}:** {amount:,.2f}")
        md_parts.append("")
    else:
        md_parts.append("No financial amounts were extracted from the documents.")
        md_parts.append("")
    
    md_parts.append("### Notes")
    md_parts.append("")
    md_parts.append("- All amounts were extracted from PDF receipts using OCR and AI analysis")
    md_parts.append("- Detailed breakdowns are available in the Excel summary file")
    md_parts.append("- Non-receipt documents are included in the analysis but may not have financial data")
    md_parts.append("")
    
    markdown_content = '\n'.join(md_parts)
    
    # If no CSS file provided, create a default one
    if not css_file:
        default_css = """
        body { font-family: Arial, sans-serif; }
        h1 { color: #333; }
        h2 { color: #666; margin-top: 20px; }
        strong { color: #444; }
        """
        temp_css = output_path.parent / "temp_style.css"
        with open(temp_css, 'w') as f:
            f.write(default_css)
        css_file = str(temp_css)
        
        # Create the PDF
        create_summary_pdf(markdown_content, str(output_path), css_file)
        
        # Clean up temp CSS
        temp_css.unlink()
    else:
        create_summary_pdf(markdown_content, str(output_path), css_file)



def create_excel_summary(directory: Path, all_data: Dict, timestamp: str = None) -> Path:
    """
    Create an Excel file with structured summary data.
    
    Args:
        directory: Root directory for the analysis
        all_data: Complete in-memory data structure
        timestamp: Run timestamp
        
    Returns:
        Path to the created Excel file
    """
    # Prepare summary data for main sheet
    summary_rows = []
    
    for relative_path, data in all_data.items():
        if data.get('status') == 'success':
            # Parse the processed date to datetime
            processed_date = data.get('processed_date', '')
            if processed_date:
                try:
                    processed_dt = datetime.fromisoformat(processed_date)
                except:
                    processed_dt = None
            else:
                processed_dt = None
            
            if data.get('aggregate', False):
                # Aggregate summary
                row = {
                    'Directory': str(relative_path),
                    'Type': 'Aggregate',
                    'Documents': data.get('total_documents', 0),
                    'Date': processed_dt.date() if processed_dt else None
                }
                # Add currency columns
                for amt_data in data.get('extracted_amounts', []):
                    currency = amt_data['currency']
                    row[f'{currency} Total'] = amt_data['amount']
            else:
                # Regular PDF directory summary
                row = {
                    'Directory': str(relative_path),
                    'Type': 'PDFs',
                    'Documents': data.get('total_files', 0),
                    'Date': processed_dt.date() if processed_dt else None
                }
                # Add currency columns from extracted amounts
                for amt_data in data.get('extracted_amounts', []):
                    currency = amt_data['currency']
                    row[f'{currency} Total'] = amt_data['amount']
            
            summary_rows.append(row)
    
    if not summary_rows:
        console.print("[yellow]No data found for Excel summary[/yellow]")
        return None
    
    # Create DataFrame for summary
    summary_df = pd.DataFrame(summary_rows)
    
    # Sort by directory path
    summary_df = summary_df.sort_values('Directory')
    
    # Use provided timestamp or generate new one
    if timestamp is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_file = directory / f"summary_{timestamp}.xlsx"
    
    # Create workbook with openpyxl directly for better control
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Write summary sheet first
    headers = list(summary_df.columns)
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Write data
    for row_idx, row_data in enumerate(summary_df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            
            # Format numbers
            if isinstance(value, (int, float)) and not pd.isna(value):
                if 'Total' in headers[col_idx-1] or 'CHF' in headers[col_idx-1]:
                    cell.number_format = '#,##0.00'
                else:
                    cell.number_format = '#,##0'
            
            # Format dates
            elif headers[col_idx-1] == 'Date' and value:
                cell.number_format = 'YYYY-MM-DD'
    
    # Create Excel table (don't include the sum row in the table)
    table_range = f"A1:{get_column_letter(len(headers))}{len(summary_df) + 1}"
    tab = ExcelTable(displayName="SummaryTable", ref=table_range)
    
    # Add table style
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    ws_summary.add_table(tab)
    
    # Freeze panes (header row)
    ws_summary.freeze_panes = ws_summary['A2']
    
    # Auto-adjust column widths for summary sheet
    for column in ws_summary.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_summary.column_dimensions[column_letter].width = adjusted_width
    
    # Add sum row at the bottom of the summary table
    last_row = len(summary_df) + 2
    ws_summary.cell(row=last_row, column=1, value="TOTAL").font = Font(bold=True)
    
    # Calculate sums for numeric columns
    for col_idx, header in enumerate(headers, 1):
        if 'Total' in header or header == 'Documents':
            # Sum the values in this column using SUBTOTAL for filtering support
            col_letter = get_column_letter(col_idx)
            sum_formula = f"=SUBTOTAL(9,{col_letter}2:{col_letter}{last_row-1})"
            cell = ws_summary.cell(row=last_row, column=col_idx, value=sum_formula)
            cell.font = Font(bold=True)
            
            # Apply number format
            if 'Total' in header or 'CHF' in header:
                cell.number_format = '#,##0.00'
            else:
                cell.number_format = '#,##0'
    
    # Save workbook
    wb.save(excel_file)
    
    return excel_file


def display_summary_table(all_data: Dict):
    """
    Display a rich table with summary data in the console.
    
    Args:
        all_data: Complete in-memory data structure
    """
    from rich.table import Table
    
    # Create the table
    table = Table(title="PDF Analysis Summary", show_lines=True)
    
    # Add columns
    table.add_column("Directory", style="cyan", no_wrap=True)
    table.add_column("Type", style="magenta")
    table.add_column("Documents", justify="right", style="green")
    table.add_column("Date", style="blue")
    
    # Track which currencies we have
    all_currencies = set()
    for data in all_data.values():
        if data.get('status') == 'success' and data.get('extracted_amounts'):
            for amt in data['extracted_amounts']:
                all_currencies.add(amt['currency'])
    
    # Add columns for each currency found
    for currency in sorted(all_currencies):
        table.add_column(f"{currency} Total", justify="right", style="yellow")
    
    # Track totals
    total_docs = 0
    currency_totals = defaultdict(float)
    
    # Add rows
    for relative_path, data in sorted(all_data.items()):
        if data.get('status') == 'success':
            # Parse the processed date to datetime
            processed_date = data.get('processed_date', '')
            date_str = ""
            if processed_date:
                try:
                    processed_dt = datetime.fromisoformat(processed_date)
                    date_str = processed_dt.strftime('%Y-%m-%d')
                except:
                    pass
            
            # Basic data
            dir_type = 'Aggregate' if data.get('aggregate', False) else 'PDFs'
            doc_count = data.get('total_documents', data.get('total_files', 0))
            total_docs += doc_count
            
            # Currency data
            currency_amounts = {curr: 0 for curr in all_currencies}
            
            if data.get('extracted_amounts'):
                for amt in data['extracted_amounts']:
                    currency_amounts[amt['currency']] = amt['amount']
                    currency_totals[amt['currency']] += amt['amount']
            
            # Build row data
            row_data = [
                str(relative_path),
                dir_type,
                str(doc_count),
                date_str
            ]
            
            # Add currency amounts
            for currency in sorted(all_currencies):
                amount = currency_amounts.get(currency, 0)
                row_data.append(f"{amount:,.2f}" if amount > 0 else "-")
            
            table.add_row(*row_data)
    
    # Add total row
    total_row = [
        "[bold]TOTAL[/bold]",
        "-",
        f"[bold]{total_docs}[/bold]",
        "-"
    ]
    
    for currency in sorted(all_currencies):
        total_row.append(f"[bold]{currency_totals[currency]:,.2f}[/bold]" if currency_totals[currency] > 0 else "-")
    
    table.add_row(*total_row)
    
    # Display the table
    console.print(table)


def analyze_pdfs_recursively(directory: Path, output_format: str = "txt", cleanup: bool = True, max_workers: int = 4) -> List[Dict]:
    """
    Recursively analyze PDFs in directories and create hierarchical summaries.
    """
    results = []
    
    # Generate a single timestamp for the entire run
    run_timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TaskProgressColumn(),
        TimeElapsedColumn(),
        console=console,
        transient=False
    ) as progress:
        # Find all directories with PDFs
        scan_task = progress.add_task("[cyan]📂 Scanning for PDF files...", total=None)
        pdf_map = find_pdfs_recursively(directory)
        progress.update(scan_task, completed=100, total=100)
        
        if not pdf_map:
            console.print("[yellow]No PDF files found in the directory tree.[/yellow]")
            return results
        
        progress.update(scan_task, description=f"[green]✅ Found {sum(len(pdfs) for pdfs in pdf_map.values())} PDFs in {len(pdf_map)} directories")
        
        task = progress.add_task(f"[cyan]📄 Processing directories (using {max_workers} threads)...", total=len(pdf_map))
        
        # Group results by parent directory
        results_by_parent = defaultdict(list)
        
        # Process directories in parallel
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            # Submit all tasks
            future_to_dir = {}
            for dir_path, pdf_files in pdf_map.items():
                future = executor.submit(process_pdf_directory, dir_path, False)
                future_to_dir[future] = (dir_path, pdf_files)
            
            # Process results as they complete
            for future in as_completed(future_to_dir):
                dir_path, pdf_files = future_to_dir[future]
                rel_path = dir_path.relative_to(directory) if dir_path != directory else Path('.')
                
                try:
                    result = future.result()
                    results.append(result)
                    
                    # Show completion for this directory
                    if result['status'] == 'success':
                        progress.update(task, description=f"[green]✓ Processed [bold]{rel_path}[/bold]: {len(pdf_files)} PDFs")
                        
                        # Track results by parent directory
                        parent = dir_path.parent
                        if parent != directory and parent in pdf_map.keys():
                            # Don't create aggregate for directories that have their own PDFs
                            pass
                        else:
                            results_by_parent[parent].append(result)
                    else:
                        progress.update(task, description=f"[yellow]⚠ Skipped [bold]{rel_path}[/bold]: no PDFs")
                        
                except Exception as e:
                    console.print(f"[red]Error processing {dir_path}: {e}[/red]")
                    results.append({
                        'directory': str(dir_path),
                        'total_files': 0,
                        'status': 'error',
                        'error': str(e)
                    })
                
                progress.advance(task)
        
        # Create aggregate summaries for parent directories
        console.print("\n[blue]Creating aggregate summaries...[/blue]")
        
        # Create a complete data structure with all results
        all_data = {}
        for result in results:
            if result['status'] == 'success':
                dir_path = Path(result['directory'])
                rel_path = str(dir_path.relative_to(directory)) if dir_path != directory else '.'
                all_data[rel_path] = result
        
        # Build hierarchy and create aggregates bottom-up
        dirs_to_process = set(Path(p) for p in all_data.keys())
        aggregated_data = {}
        
        # Sort paths by depth (deepest first)
        sorted_paths = sorted(dirs_to_process, key=lambda x: len(x.parts), reverse=True)
        
        # Process each directory and create aggregates for parents
        for dir_path in sorted_paths:
            rel_path = str(dir_path) if str(dir_path) != '.' else '.'
            
            # Find immediate children
            children = []
            for other_path in all_data.keys():
                other = Path(other_path) if other_path != '.' else Path('.')
                if other.parent == dir_path and other != dir_path:
                    children.append(other_path)
            
            # Also check aggregated data for children
            for other_path in aggregated_data.keys():
                other = Path(other_path) if other_path != '.' else Path('.')
                if other.parent == dir_path and other != dir_path and other_path not in children:
                    children.append(other_path)
            
            # If this directory has children, create an aggregate
            if children and rel_path not in all_data:
                child_results = []
                for child in children:
                    if child in all_data:
                        child_results.append(all_data[child])
                    elif child in aggregated_data:
                        child_results.append(aggregated_data[child])
                
                if child_results:
                    aggregate_data = create_aggregate_summary(directory / dir_path, child_results)
                    aggregated_data[rel_path] = {
                        'directory': str(directory / dir_path),
                        'processed_date': datetime.now().isoformat(),
                        'aggregate': True,
                        'subdirectories': aggregate_data['subdirectories'],
                        'total_documents': aggregate_data['total_documents'],
                        'extracted_amounts': aggregate_data['extracted_amounts'],
                        'summary': aggregate_data['summary_text'],
                        'status': 'success'
                    }
        
        # Merge all data
        all_data.update(aggregated_data)
    
    # Create master JSON with all data
    master_json = create_master_json(directory, all_data, run_timestamp)
    
    # Create Excel summary at root level
    excel_file = create_excel_summary(directory, all_data, run_timestamp)
    
    # Display summary table in console
    console.print("\n[bold blue]Summary Results:[/bold blue]")
    display_summary_table(all_data)
    
    # Create styled PDF summary
    summary_pdf = directory / f"summary_{run_timestamp}.pdf"
    
    # Try to use the same CSS file as image summaries if available
    # First check script directory
    script_dir = Path(__file__).parent
    css_file = script_dir / "styles.css"
    if not css_file.exists():
        # Then check output directory
        css_file = directory / "styles.css"
        if not css_file.exists():
            css_file = None
    
    create_summary_pdf_from_data(all_data, summary_pdf, str(css_file) if css_file else None)
    
    # Check for rate limiting in the summaries
    rate_limited = 0
    for data in all_data.values():
        if data.get('status') == 'success' and 'rate limiting' in data.get('summary', '').lower():
            rate_limited += 1
    
    if rate_limited > 0:
        console.print(f"\n[yellow]⚠️  Warning: {rate_limited} summaries were affected by API rate limiting.[/yellow]")
        console.print("[yellow]   The service is currently overloaded. Consider re-running later for complete summaries.[/yellow]")
    
    # Print final summary
    console.print(f"\n[green]✅ Analysis complete! Files created:[/green]")
    console.print(f"   • summary_{run_timestamp}.json")
    console.print(f"   • summary_{run_timestamp}.xlsx")
    console.print(f"   • summary_{run_timestamp}.pdf")
    
    return results


def detect_processing_mode(directory: Path = None) -> str:
    """
    Detect whether to use image processing or PDF recursive processing.
    
    Returns:
        'images' for image processing (PNG/JPG/JPEG in current dir)
        'pdfs' for recursive PDF processing (PDFs in subdirectories)
        'none' if no processable files found
    """
    if directory is None:
        directory = Path.cwd()
    
    # Check for images in current directory
    image_patterns = ['*.png', '*.jpg', '*.jpeg', '*.PNG', '*.JPG', '*.JPEG']
    image_files = []
    for pattern in image_patterns:
        image_files.extend(directory.glob(pattern))
    
    if image_files:
        return 'images'
    
    # Check for PDFs in subdirectories
    pdf_map = find_pdfs_recursively(directory)
    if pdf_map:
        return 'pdfs'
    
    return 'none'


def sort_by_directory_order(files):
    """
    Sort files to match the exact order shown in the directory listing.

    Args:
        files (List[str]): List of file paths.

    Returns:
        List[str]: Sorted list of file paths.
    """
    # Get the full path of each file
    full_paths = [os.path.abspath(f) for f in files]

    # Sort based on the order in which files appear in the directory
    return sorted(full_paths, key=lambda x: os.path.getmtime(x))


def crop_image(image_path, padding=10):
    """
    Crop an image to remove any surrounding whitespace and add padding for print borders.

    Args:
        image_path (str): Path to the image file.
        padding (int): Amount of padding to add around the cropped area. Default is 10 pixels.

    Returns:
        PIL.Image.Image: Cropped image with padding.
    """
    img = Image.open(image_path)
    if img.mode == 'RGBA':
        img = img.convert('RGB')
    bg = Image.new(img.mode, img.size, img.getpixel((0, 0)))
    diff = ImageChops.difference(img, bg)
    diff = ImageChops.add(diff, diff, 2.0, -100)
    bbox = diff.getbbox()

    if bbox:
        left, upper, right, lower = bbox
        cropped_img = img.crop((left, upper, right, lower))

        # Create a new image with horizontal padding
        padded_img = Image.new(img.mode, (cropped_img.width + 2 * padding, cropped_img.height), img.getpixel((0, 0)))
        padded_img.paste(cropped_img, (padding, 0))
        return padded_img
    else:
        return img


def convert_images_to_pdf(image_paths, output_path, padding_ratio=0.02):
    page_width, page_height = landscape(A4)
    padding = int(page_width * padding_ratio)
    c = canvas.Canvas(output_path, pagesize=(page_width, page_height))

    for image_path in image_paths:
        img = crop_image(image_path, padding)
        img_width, img_height = img.size

        max_width = page_width - 2 * padding
        max_height = page_height
        scale = min(max_width / img_width, max_height / img_height)

        scaled_width = int(img_width * scale)
        scaled_height = int(img_height * scale)

        x_position = (page_width - scaled_width) / 2
        y_position = (page_height - scaled_height) / 2

        img_buffer = io.BytesIO()
        # img.save(img_buffer, format='JPEG', quality=85, optimize=True)
        img.save(img_buffer, format='PNG', quality=100, optimize=True)
        img_buffer.seek(0)

        c.drawImage(ImageReader(img_buffer), x_position, y_position, width=scaled_width, height=scaled_height)
        c.showPage()

    c.save()


def compress_pdf(input_path, output_path):
    reader = PdfReader(input_path)
    writer = PdfWriter()

    for page in reader.pages:
        page.compress_content_streams()  # This is CPU intensive!
        writer.add_page(page)

    for key, value in reader.metadata.items():
        writer.add_metadata({key: value})

    with open(output_path, "wb") as f:
        writer.write(f)


def convert_images_to_pdf(image_paths, output_path, padding_ratio=0.02):
    page_width, page_height = landscape(A4)
    padding = int(page_width * padding_ratio)
    c = canvas.Canvas(output_path, pagesize=(page_width, page_height))
    c.setPageCompression(1)  # Enable compression

    for image_path in image_paths:
        img = crop_image(image_path, padding)
        img_width, img_height = img.size

        max_width = page_width - 2 * padding
        max_height = page_height
        scale = min(max_width / img_width, max_height / img_height)

        scaled_width = int(img_width * scale)
        scaled_height = int(img_height * scale)

        x_position = (page_width - scaled_width) / 2
        y_position = (page_height - scaled_height) / 2

        img_buffer = io.BytesIO()
        # Convert RGBA to RGB before saving as JPEG
        if img.mode == 'RGBA':
            img = img.convert('RGB')

        # Note: JPEG compression is lossy, so use PNG for better quality
        # This here makes all the difference.
        # img.save(img_buffer, format='JPEG', quality=85, optimize=True)
        img.save(img_buffer, format='PNG', quality=100, optimize=True)
        img_buffer.seek(0)

        c.drawImage(ImageReader(img_buffer), x_position, y_position, width=scaled_width, height=scaled_height)
        c.showPage()

    c.save()


def improve_image(image):
    # Increase contrast
    enhancer = ImageEnhance.Contrast(image)
    image = enhancer.enhance(1.5)

    # Increase sharpness
    enhancer = ImageEnhance.Sharpness(image)
    image = enhancer.enhance(1.5)

    return image


def apply_ocr(pdf_path):
    try:
        with SuppressOutput():
            images = convert_from_path(pdf_path)

        extracted_text = ""
        doc = fitz.open()

        for image in images:
            image = improve_image(image)

            ocr_result = pytesseract.image_to_data(image, lang='eng+pol+deu+fra+spa', output_type=pytesseract.Output.DICT)

            img_bytes = io.BytesIO()
            image.save(img_bytes, format='PNG', quality=100, optimize=True)  # Use JPEG with lower quality
            img_bytes.seek(0)
            pix = fitz.Pixmap(img_bytes)
            page = doc.new_page(width=pix.width, height=pix.height)

            page.insert_image(fitz.Rect(0, 0, pix.width, pix.height), pixmap=pix)

            confidence_threshold = 60
            for j in range(len(ocr_result['text'])):
                if ocr_result['text'][j].strip() and int(ocr_result['conf'][j]) > confidence_threshold:
                    x, y, w, h = ocr_result['left'][j], ocr_result['top'][j], ocr_result['width'][j], ocr_result['height'][j]
                    text = ocr_result['text'][j]
                    extracted_text += text + " "

                    try:
                        page.insert_textbox(
                            fitz.Rect(x, y, x+w, y+h),
                            text,
                            fontname="helv",
                            fontsize=h,
                            color=(0, 0, 0, 0)
                        )
                    except Exception:
                        pass

            extracted_text += "\n"

        doc.save(pdf_path, deflate=True, garbage=4, clean=True)  # Use more aggressive compression
        doc.close()

        return extracted_text.strip()

    except Exception as e:
        print(f"Error in apply_ocr: {str(e)}")
        return ""


def are_images_similar(img1_path, img2_path, threshold=0.90):
    """
    Check if two images are similar using structural similarity index.

    Args:
        img1_path (str): Path to the first image.
        img2_path (str): Path to the second image.
        threshold (float): Similarity threshold.

    Returns:
        bool: True if images are similar, False otherwise.
    """
    img1 = Image.open(img1_path).convert('L')
    img2 = Image.open(img2_path).convert('L')
    img2 = img2.resize(img1.size)
    img1_array = np.array(img1)
    img2_array = np.array(img2)
    similarity = ssim(img1_array, img2_array)
    return similarity > threshold


def check_similarity(args):
    img, image, threshold = args
    return are_images_similar(img, image, threshold)

def filter_similar_images(image_files, threshold=0.95, num_workers=None, progress=None, filter_task=None):
    """
    Filter out similar images from a list of image files using parallel processing.

    Args:
        image_files (List[str]): List of image file paths.
        threshold (float): Similarity threshold.
        num_workers (int, optional): Number of parallel workers to use. Default is None, which uses all available cores.

    Returns:
        List[str]: Filtered list of image file paths.
    """
    if not image_files:
        return []

    filtered_images = [image_files[0]]
    remaining_images = image_files[1:]

    with ProcessPoolExecutor(max_workers=num_workers) as executor:
        for image in remaining_images:
            tasks = [(img, image, threshold) for img in filtered_images]
            similarities = list(executor.map(check_similarity, tasks))
            if not any(similarities):
                filtered_images.append(image)
            if progress:
                progress.advance(filter_task)

    return filtered_images


def process_image(args):
    """
    Process a single image file.

    Args:
        args (tuple): Tuple containing (png_file, temp_dir, i, use_original_resolution).

    Returns:
        tuple: (pdf_path, extracted_text)
    """
    png_file, temp_dir, i = args
    pdf_path = os.path.join(temp_dir, f"temp_{i}.pdf")
    try:
        with SuppressOutput():
            convert_images_to_pdf([png_file], pdf_path)  # Make sure this line is correct
            extracted_text = apply_ocr(pdf_path)
        return pdf_path, extracted_text
    except Exception as e:
        console.print(f"[red]Error processing image {png_file}: {str(e)}[/red]")
        console.print(f"[yellow]Stack trace:[/yellow]\n{traceback.format_exc()}")
        return None, None


def clean_text(text):
    """
    Clean and preprocess text by removing stop words and meaningless sentences.

    Args:
        text (str): Input text to clean.

    Returns:
        str: Cleaned text.
    """
    nlp = spacy.load("en_core_web_sm")
    stop_words = set(stopwords.words('english'))
    sentences = sent_tokenize(text)
    meaningful_sentences = []
    for sentence in sentences:
        words = word_tokenize(sentence)
        filtered_words = [word for word in words if word.isalpha() and word.lower() not in stop_words]
        filtered_sentence = ' '.join(filtered_words)
        doc = nlp(filtered_sentence)
        if any(token.pos_ in ['NOUN', 'VERB', 'ADJ', 'ADV'] for token in doc):
            meaningful_sentences.append(filtered_sentence)
    cleaned_text = ' '.join(meaningful_sentences)
    return cleaned_text.strip()


def extract_text_from_pdf(pdf_path):
    """
    Extract text from a PDF file.

    Args:
        pdf_path (str): Path to the PDF file.

    Returns:
        str: Extracted text from the PDF.
    """
    doc = fitz.open(pdf_path)
    text = ""
    for page in doc:
        text += page.get_text()
    return text


def generate_summary(text, max_tokens=1000):
    """
    Generate a summary of the given text using an AI model.

    Args:
        text (str): Input text to summarize.
        max_tokens (int): Maximum number of tokens for the summary.

    Returns:
        str: Generated summary.
    """
    text = clean_text(text)

    ai_client = AIClient(config_provider=GitConfig())
    console.print(f"[blue]Using {ai_client.name} to generate a summary...[/blue]")
    prompt = f"""
    Please provide a concise summary of the following text, followed by a bulleted list
    of the main topics and their explanations. Format the output exactly as follows,
    using markdown formatting:

    # Summary:
    [Your concise summary here]

    # Main Topics:

    ## [Topic 1]:

    [Your Summary of Topic 1: you can be quite a bit more detailed here and use
    multiple sentences if necessary and give context]

    - [Explanation point 1]
    - [Explanation point 2]

    ## [Topic 2]:

    [Your Summary of Topic 2: you can be quite a bit more detailed here and use
    multiple sentences if necessary and give context]

    - [Explanation point 1]
    - [Explanation point 2]

    [and so on...]

    Important: Start your response directly with the summary. Do not use
    any introductory phrases like "Sure," "Here's," or "Certainly."

    You don't need to specially mention things like "Multilingual elements
    suggesting the document may be available in multiple languages".

    If you see a Microsoft Teams or other video conferincing app window in the
    screenshots, you do not need to describe that interface. You can assume the
    reader knows what it looks like. I absolutely do not want you to give a
    description of the teams interface. so no "Meeting controls (e.g., Leave, Take control, Chat, Raise hand)"
    etc. Just focus on the content of the document.

    Here's the text to summarize:

    {text}
    """

    try:
        response = ai_client.prompt(prompt, tokens=max_tokens)
        return response
    except Exception as e:
        console.print(f"[red]Error getting response from AI: {str(e)}[/red]")
        console.print(f"[yellow]Prompt:[/yellow] {prompt}")
        return None


def parse_css_value(value, default):
    """
    Parse CSS value and return the corresponding numeric value.

    Args:
        value (str): CSS value to parse.
        default: Default value to return if parsing fails.

    Returns:
        int or str: Parsed value.
    """
    if isinstance(value, str):
        if value.endswith('%'):
            return value
        match = re.match(r'(\d+)(?:px)?', value)
        if match:
            return int(match.group(1))
    return default


def parse_color(color_string):
    """
    Parse color string and return the corresponding Color object.

    Args:
        color_string (str): Color string to parse.

    Returns:
        reportlab.lib.colors.Color: Parsed color object.
    """
    if not color_string:
        return black

    color_string = color_string.strip()
    if color_string.startswith('#'):
        if len(color_string) == 4:
            r = int(color_string[1] * 2, 16) / 255.0
            g = int(color_string[2] * 2, 16) / 255.0
            b = int(color_string[3] * 2, 16) / 255.0
            return Color(r, g, b)
        elif len(color_string) == 7:
            try:
                r = int(color_string[1:3], 16) / 255.0
                g = int(color_string[3:5], 16) / 255.0
                b = int(color_string[5:7], 16) / 255.0
                return Color(r, g, b)
            except ValueError:
                print(f"Invalid hex color: {color_string}")
                return black
        else:
            print(f"Invalid hex color format: {color_string}")
            return black
    else:
        try:
            return Color(color_string)
        except:
            print(f"Invalid color name: {color_string}")
            return black


def create_summary_pdf(summary, output_path, css_path):
    """
    Create a styled PDF summary from the given text and CSS file.

    Args:
        summary     (str): Summary text in markdown format.
        output_path (str): Path to save the output PDF.
        css_path    (str): Path to the CSS file for styling.
    """
    if not summary:
        raise ValueError("Summary content is empty")

    doc = SimpleDocTemplate(output_path, pagesize=A4, rightMargin=inch, leftMargin=inch, topMargin=inch, bottomMargin=inch)

    # Get sample stylesheet as a fallback
    sample_styles = getSampleStyleSheet()

    # Parse CSS file
    with open(css_path, 'r') as css_file:
        css = cssutils.parseString(css_file.read())

    # Create styles based on CSS
    styles = {}
    for rule in css:
        if rule.type == rule.STYLE_RULE:
            selector = rule.selectorText.strip('.')
            style_dict = {p.name: p.value for p in rule.style}

            # Use sample style as a base, then override with CSS values
            base_style = sample_styles['BodyText']
            new_style = ParagraphStyle(
                selector,
                parent=base_style,
                fontSize       = parse_css_value(style_dict.get('font-size'),     base_style.fontSize),
                leading        = parse_css_value(style_dict.get('line-height'),   base_style.leading),
                spaceBefore    = parse_css_value(style_dict.get('margin-top'),    base_style.spaceBefore),
                spaceAfter     = parse_css_value(style_dict.get('margin-bottom'), base_style.spaceAfter),
                leftIndent     = parse_css_value(style_dict.get('padding-left'),  base_style.leftIndent),
                rightIndent    = parse_css_value(style_dict.get('padding-right'), base_style.rightIndent),
                firstLineIndent= parse_css_value(style_dict.get('text-indent'),   base_style.firstLineIndent),
                textColor      = parse_color    (style_dict.get('color',            '#000000')),
                backColor      = parse_color    (style_dict.get('background-color', '#FFFFFF')),
            )

            # Handle text alignment
            text_align = style_dict.get('text-align', '').lower()
            if text_align == 'justify':
                new_style.alignment = TA_JUSTIFY
            elif text_align == 'center':
                new_style.alignment = TA_CENTER
            elif text_align == 'right':
                new_style.alignment = TA_RIGHT
            else:
                new_style.alignment = TA_LEFT

            # Handle text transformation
            if style_dict.get('text-transform') == 'uppercase':
                new_style.textTransform = 'uppercase'

            styles[selector] = new_style

            # print(f"Selector: {selector}")
            # print(f"Text Color: {new_style.textColor}")
            # print(f"Background Color: {new_style.backColor}")

    story = []

    # Convert Markdown to HTML
    html = markdown.markdown(summary)
    soup = BeautifulSoup(html, 'html.parser')

    # Get the available width for content
    available_width = A4[0] - 2*inch  # Subtracting left and right margins

    # Parse HTML and create paragraphs or tables for styled elements
    for element in soup.descendants:
        if element.name in ['h1', 'h2']:
            #if element.name == 'h2':
            story.append(Spacer(1, 12))
            style_name = 'heading1' if element.name == 'h1' else 'heading2'
            text = element.get_text().strip()
            style = styles.get(style_name, sample_styles['Heading1'])

            # Apply text transformation if specified in CSS
            if hasattr(style, 'textTransform') and style.textTransform == 'uppercase':
                text = text.upper()

            # Create a table for both heading1 and heading2 with background color and border
            t = Table([[text]], colWidths=[available_width])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,-1), style.backColor),
                ('TEXTCOLOR', (0,0), (-1,-1), style.textColor),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('FONTNAME', (0,0), (-1,-1), style.fontName),
                ('FONTSIZE', (0,0), (-1,-1), style.fontSize),
                ('BOTTOMPADDING', (0,0), (-1,-1), parse_css_value(style_dict.get('padding-bottom', '5'), 5)),
                ('TOPPADDING', (0,0), (-1,-1), parse_css_value(style_dict.get('padding-top', '5'), 5)),
                ('LEFTPADDING', (0,0), (-1,-1), parse_css_value(style_dict.get('padding-left', '10'), 10)),
                ('RIGHTPADDING', (0,0), (-1,-1), parse_css_value(style_dict.get('padding-right', '10'), 10)),
                ('LEADING', (0,0), (-1,-1), style.leading),
            ]))
            story.append(t)
            if element.name == 'h1':
                story.append(Spacer(1, parse_css_value(style.spaceAfter, 12)))
        elif element.name == 'p':
            story.append(Paragraph(element.get_text(), styles.get('bodytext', sample_styles['BodyText'])))
        elif element.name == 'ul':
            for li in element.find_all('li'):
                story.append(Paragraph(f'• {li.get_text()}', styles.get('bullet', sample_styles['BodyText'])))
                story.append(Spacer(1, 6))

    # Add a page break for multi-page support
    story.append(PageBreak())

    # Build PDF
    doc.build(story)


@app.command()
def summarize(
    input_files:      List[str] = typer.Argument(None,                               help="Paths to input files (PDFs or images)"),
    output_file:            str = typer.Option("summary.pdf", "-o", "--output-file", help="Path to the output summary PDF file"),
    css_file:     Optional[str] = typer.Option(None,          "-c", "--css-file",    help="Path to the CSS file for styling"),
    max_tokens:             int = typer.Option(1000,          "-m", "--max-tokens",  help="Maximum number of tokens for summary generation", show_default=True),
    similarity_threshold: float = typer.Option(0.95,          "-s", "--similarity",  help="Threshold for image similarity (0-1). Use 1 for no similarity check.", show_default=True),
    no_ai:                 bool = typer.Option(False,         "-n", "--no-ai",       help="Skip AI recognition and summary generation"),
):
    """
    Generate a summary from input files (PDFs or images).

    This command automatically detects the type of processing needed:
    - If PNG/JPG/JPEG files are found in the current directory: Creates a styled PDF summary
    - If PDFs are found in subdirectories: Creates recursive text summaries with totals
    """
    console.print("[blue]Starting summarize command...[/blue]")

    # Check if a single directory path was provided
    if input_files and len(input_files) == 1 and os.path.isdir(input_files[0]):
        directory = Path(input_files[0])
        mode = detect_processing_mode(directory)
        
        if mode == 'pdfs':
            console.print(f"[blue]Found PDFs in subdirectories of {directory}. Processing recursively for text summaries...[/blue]")
            
            # Process PDFs recursively
            results = analyze_pdfs_recursively(directory, output_format="txt")
            
            # Summary report
            console.print("\n[bold]Processing Summary:[/bold]")
            total_pdfs = sum(r.get('total_files', r.get('pdfs', 0)) for r in results if r['status'] == 'success')
            successful_dirs = sum(1 for r in results if r['status'] == 'success')
            failed_dirs = sum(1 for r in results if r['status'] == 'error')
            
            console.print(f"  Directories processed: {successful_dirs + failed_dirs}")
            console.print(f"  Total PDFs analyzed: {total_pdfs}")
            console.print(f"  Successful summaries: {successful_dirs}")
            
            if failed_dirs > 0:
                console.print(f"  [red]Failed summaries: {failed_dirs}[/red]")
                
            # Check for rate limiting issues
            if any('rate limiting' in r.get('summary', '').lower() for r in results if r.get('status') == 'success'):
                console.print("\n[yellow]⚠️  Note: Some summaries were affected by API rate limiting.[/yellow]")
                console.print("[yellow]   You may want to re-run the analysis later when the service is less busy.[/yellow]")
            
            # Exit early for PDF processing
            return
        else:
            # Try to process as directory with images
            os.chdir(directory)
            input_files = None  # Reset to trigger detection below

    # Detect processing mode if no explicit input files provided
    if input_files is None or len(input_files) == 0:
        mode = detect_processing_mode()
        
        if mode == 'images':
            console.print("[blue]Found image files. Processing as screenshots for PDF summary...[/blue]")
            all_files = glob.glob("*.png") + glob.glob("*.jpg") + glob.glob("*.jpeg")
            input_files = sort_by_directory_order(all_files)
            # Continue with existing image processing logic below
            
        elif mode == 'pdfs':
            console.print("[blue]Found PDFs in subdirectories. Processing recursively for text summaries...[/blue]")
            
            # Process PDFs recursively
            results = analyze_pdfs_recursively(Path.cwd(), output_format="txt")
            
            # Summary report
            console.print("\n[bold]Processing Summary:[/bold]")
            total_pdfs = sum(r.get('total_files', r.get('pdfs', 0)) for r in results if r['status'] == 'success')
            successful_dirs = sum(1 for r in results if r['status'] == 'success')
            failed_dirs = sum(1 for r in results if r['status'] == 'error')
            
            console.print(f"  Directories processed: {successful_dirs + failed_dirs}")
            console.print(f"  Total PDFs analyzed: {total_pdfs}")
            console.print(f"  Successful summaries: {successful_dirs}")
            
            if failed_dirs > 0:
                console.print(f"  [red]Failed summaries: {failed_dirs}[/red]")
                
            # Check for rate limiting issues
            if any('rate limiting' in r.get('summary', '').lower() for r in results if r.get('status') == 'success'):
                console.print("\n[yellow]⚠️  Note: Some summaries were affected by API rate limiting.[/yellow]")
                console.print("[yellow]   You may want to re-run the analysis later when the service is less busy.[/yellow]")
            
            # Exit early for PDF processing
            return
            
        else:
            console.print("[red]No processable files found in the current directory or subdirectories.[/red]")
            raise typer.Exit(1)

    # Use default CSS file if no CSS file is provided
    if css_file is None:
        script_dir = os.path.dirname(os.path.realpath(__file__))
        css_file = os.path.join(script_dir, "styles.css")

    pdf_files = [f for f in input_files if f.lower().endswith('.pdf')]
    image_files = [f for f in input_files if f.lower().endswith(('.png', '.jpg', '.jpeg'))]

    all_text = ""

    with Progress() as progress:
        try:
            merger = PdfMerger()

            if pdf_files:
                pdf_task = progress.add_task("[green]Processing PDFs...", total=len(pdf_files))
                for pdf in pdf_files:
                    merger.append(pdf)
                    text = extract_text_from_pdf(pdf)
                    all_text += text + "\n"
                    progress.advance(pdf_task)

            filtered_files = []
            if image_files:
                valid_images = []
                for image in image_files:
                    try:
                        with Image.open(image) as img:
                            valid_images.append(image)
                    except UnidentifiedImageError:
                        console.print(f"[red]Skipping invalid image file: {image}[/red]")

                sorted_files = sort_by_directory_order(valid_images)

                # Skip similarity check if threshold is 1
                if similarity_threshold == 1:
                    filtered_files = sorted_files
                    console.print("[blue]Skipping image similarity check as threshold is set to 1.[/blue]")
                else:
                    # Filter similar images with progress update
                    filter_task = progress.add_task("[blue]Filtering images...", total=len(sorted_files))
                    filtered_files = filter_similar_images(sorted_files, threshold=similarity_threshold, num_workers=multiprocessing.cpu_count(), progress=progress, filter_task=filter_task)
                    progress.update(filter_task, completed=len(sorted_files))

                image_task = progress.add_task("[blue]Processing images...", total=len(filtered_files))

                with tempfile.TemporaryDirectory() as temp_dir:
                    num_cores = multiprocessing.cpu_count()
                    with ProcessPoolExecutor(max_workers=num_cores) as executor:
                        futures = {executor.submit(process_image, (png, temp_dir, i)): i for i, png in enumerate(filtered_files)}
                        results = []
                        for future in as_completed(futures):
                            try:
                                i = futures[future]
                                pdf_path, extracted_text = future.result()
                                if pdf_path and extracted_text:
                                    results.append((pdf_path, extracted_text))
                                progress.advance(image_task)
                            except Exception as e:
                                console.print(f"[red]Error processing image: {e}[/red]")
                                console.print(f"[yellow]Stack trace:[/yellow]\n{traceback.format_exc()}")

                    for pdf_path, extracted_text in results:
                        merger.append(pdf_path)
                        all_text += extracted_text + "\n"

            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_output_file:
                temp_output_pdf_path = temp_output_file.name

            merger.write(temp_output_pdf_path)
            merger.close()

        except Exception as e:
            console.print(f"[red]Error creating the final PDF: {str(e)}[/red]")
            console.print(f"[yellow]Stack trace:[/yellow]\n{traceback.format_exc()}")
            return

        progress.stop()

        # Generate final output filename
        if image_files and filtered_files:
            first_file_modification_time = datetime.fromtimestamp(os.path.getmtime(filtered_files[0]))
        elif pdf_files:
            first_file_modification_time = datetime.fromtimestamp(os.path.getmtime(pdf_files[0]))
        else:
            first_file_modification_time = datetime.now()

        date_time_str = first_file_modification_time.strftime("%Y%m%d_%H%M")
        output_base_name = os.path.splitext(output_file)[0]
        final_output_filename = f"{output_base_name}_{date_time_str}.pdf"
        final_output_path = os.path.abspath(final_output_filename)

        if not no_ai:
            summary = generate_summary(all_text, max_tokens)

            if summary is None:
                console.print("[red]Failed to generate summary. Please check your AI configuration and try again.[/red]")
                return

            try:
                with tempfile.TemporaryDirectory() as temp_dir:
                    summary_pdf_path = os.path.join(temp_dir, "summary.pdf")
                    create_summary_pdf(summary, summary_pdf_path, css_file)

                    final_merger = PdfMerger()
                    final_merger.append(summary_pdf_path)
                    final_merger.append(temp_output_pdf_path)

                    with open(final_output_path, "wb") as f:
                        final_merger.write(f)
                    final_merger.close()

                    # Compress the final PDF
                    compress_pdf(final_output_path, final_output_path)

                    console.print(f"[green]Compressed summary PDF created at {final_output_path}[/green]")
            except Exception as e:
                console.print(f"[red]Error creating summary PDF: {str(e)}[/red]")
                console.print(f"[yellow]Stack trace:[/yellow]\n{traceback.format_exc()}")
            finally:
                if os.path.exists(temp_output_pdf_path):
                    os.remove(temp_output_pdf_path)
        else:
            try:
                shutil.move(temp_output_pdf_path, final_output_path)

                # Compress the PDF even if no AI summary is generated
                compress_pdf(final_output_path, final_output_path)

                console.print(f"[green]Compressed PDF created without AI summary at {final_output_path}[/green]")
            except Exception as e:
                console.print(f"[red]Error moving or compressing the PDF file: {str(e)}[/red]")
                console.print(f"[yellow]Stack trace:[/yellow]\n{traceback.format_exc()}")


@app.command()
def config(
    provider:     Optional[str] = typer.Argument(None,                          help="The AI provider to configure"),
    set_default:           bool = typer  .Option(False,  "-d", "--set-default", help="Set the specified provider as default")
):
    """
    Configure, update, create, delete, clone an AI provider interactively, or set the default provider.

    For reference, here is a typical configuration for GPT and Claude:

    ```ini
    [openai]
    name = openai
    aiprovider = true
    apikey = sk-proj-...
    model = gpt-4o
    url = https://api.openai.com/v1/chat/completions
    header = {Authorization: Bearer {api_key}}
    response = response.json()['choices'][0]['message']['content']

    [claude]
    name = Claude
    aiprovider = true
    apikey = sk-ant-...
    model = claude-3-5-sonnet-20240620
    url = https://api.anthropic.com/v1/messages
    header = {x-api-key: {api_key}, anthropic-version: 2023-06-01}
    response = response.json()['content'][0]['text']
    ```

    """
    git_config = GitConfig()
    available_providers = git_config.get_available_providers()

    if set_default:
        if provider:
            if provider in available_providers:
                git_config.set_default_provider(provider)
                console.print(f"[green]Set {provider} as the default AI provider.[/green]")
            else:
                console.print(f"[red]Provider '{provider}' not found. Cannot set as default.[/red]")
        else:
            if available_providers:
                default_provider = inquirer.select(
                    message="Select the default AI provider:",
                    choices=available_providers
                ).execute()
                git_config.set_default_provider(default_provider)
                console.print(f"[green]Set {default_provider} as the default AI provider.[/green]")
            else:
                console.print("[red]No AI providers found. Cannot set a default provider.[/red]")
        return

    if not available_providers and provider is None:
        console.print("[yellow]No AI providers found. Let's create one.[/yellow]")
        provider = inquirer.text(message="Enter a name for the new provider:").execute()

    if provider is None:
        choices = available_providers + ["Create new provider", "Clone existing provider", "Set default provider"]

        selected = inquirer.select(
            message="Select an action:",
            choices=choices
        ).execute()

        if selected == "Create new provider":
            provider = inquirer.text(message="Enter the name for the new provider:").execute()
        elif selected == "Clone existing provider":
            source_provider = inquirer.select(
                message="Select a provider to clone:",
                choices=available_providers
            ).execute()
            target_provider = inquirer.text(message="Enter the name for the cloned provider:").execute()
            if git_config.clone_provider(source_provider, target_provider):
                provider = target_provider
            else:
                return
        elif selected == "Set default provider":
            default_provider = inquirer.select(
                message="Select the default AI provider:",
                choices=available_providers
            ).execute()
            git_config.set_default_provider(default_provider)
            console.print(f"[green]Set {default_provider} as the default AI provider.[/green]")
            return
        else:
            provider = selected

    git_config.configure_provider(provider)

    # Ask if the user wants to set this provider as default
    if inquirer.confirm(message=f"Do you want to set {provider} as the default AI provider?", default=False).execute():
        git_config.set_default_provider(provider)
        console.print(f"[green]Set {provider} as the default AI provider.[/green]")


@app.command()
def doc(
    ctx: typer.Context,
    title: str = typer.Option(None,  help="The title of the document"),
    toc:  bool = typer.Option(False, help="Whether to create a table of contents"),
) -> None:
    """
    Re-create the documentation and write it to the output file.

    This command generates documentation for the script, including an optional
    table of contents and custom title.
    """
    result = DocGenerator.generate_doc(__file__, title, toc)
    print(result)


#
# Main function
#
if __name__ == "__main__":
    import sys
    from typer.main import get_command,get_command_name
    if len(sys.argv) == 1 or sys.argv[1] not in [get_command_name(key) for key in get_command(app).commands.keys()]:
        sys.argv.insert(1, "summarize")
    app()
